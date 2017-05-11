import logging
import os

from openpyxl import Workbook

logging.basicConfig(format='%(levelname)s:%(asctime)s:%(message)s  \t', level=logging.INFO)

EXCEL_DIR = '/home/lucasx/PycharmProjects/DataHouse/DataSet/'


def write_excel(list_, filename):
    mkdirs_if_not_exists(EXCEL_DIR)
    wb = Workbook()
    ws = wb.active
    ws.title = "HouseInfo"
    ws.cell(row=1, column=1).value = 'address'
    ws.cell(row=1, column=2).value = 'area'
    ws.cell(row=1, column=3).value = 'block'
    ws.cell(row=1, column=4).value = 'buildYear'
    ws.cell(row=1, column=5).value = 'image'
    ws.cell(row=1, column=6).value = 'midPrice'
    ws.cell(row=1, column=7).value = 'name'
    ws.cell(row=1, column=8).value = 'saleNum'
    ws.cell(row=1, column=9).value = 'url'

    rownum = 2
    for each_item in list_:
        ws.cell(row=rownum, column=1).value = each_item.address
        ws.cell(row=rownum, column=2).value = each_item.area
        ws.cell(row=rownum, column=3).value = each_item.block
        ws.cell(row=rownum, column=4).value = each_item.buildYear
        ws.cell(row=rownum, column=5).value = each_item.image
        ws.cell(row=rownum, column=6).value = each_item.midPrice
        ws.cell(row=rownum, column=7).value = each_item.name
        ws.cell(row=rownum, column=8).value = each_item.saleNum
        ws.cell(row=rownum, column=9).value = each_item.url
        rownum += 1
    wb.save(EXCEL_DIR + filename + '.xlsx')
    logging.info('Excel生成成功!')


def mkdirs_if_not_exists(directory_):
    """create a new folder if it does not exist"""
    if not os.path.exists(directory_) or not os.path.isdir(directory_):
        os.makedirs(directory_)
