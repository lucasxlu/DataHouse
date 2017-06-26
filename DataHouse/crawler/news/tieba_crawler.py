"""
a spider for baidu tieba
"""
import os

import requests
from bs4 import BeautifulSoup
import urllib.parse

from openpyxl import Workbook


class BaiduTieba:
    def __init__(self, href="", reply="", title="", author="", createtime=""):
        self.href = href
        self.reply = reply
        self.title = title
        self.author = author
        self.createtime = createtime

    def __str__(self):
        return '{href = ' + self.href + '; reply = ' + self.reply + '; title = ' + self.title + '; title = ' \
               + self.author + '; article = ' + self.author + '; createtime = ' + self.createtime + '}';

    def __repr__(self):
        return self.__str__()


def crawl(keyword):
    pn = 1
    href_prefix = 'http://tieba.baidu.com'
    baidutieba_list = []

    while pn <= 100:
        request_url = 'http://tieba.baidu.com/f?kw=' + urllib.parse.quote(keyword) + '&ie=utf-8&pn=' + str(pn)
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Host': 'tieba.baidu.com',
            'Referer': 'http://tieba.baidu.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.106 Safari/537.36',
            'Upgrade-Insecure-Requests': '1'
        }

        cookies = dict(
            cookies_are='BAIDUID=B857A790F1F81EF15DA7C853082769A3:FG=1; BIDUPSID=B857A790F1F81EF15DA7C853082769A3; PSTM=1470617065; TIEBAUID=528f9f1fd56634f5d83cc04c; TIEBA_USERTYPE=9be462d3b2b5fb7f78ecef0e; MCITY=-218%3A; plus_cv=1::m:edb42202_0::m:1-nav:7a9d2f5d-lbs:d48613ff; H_WISE_SIDS=106846_100617_100043_107959_100102_108372_100097_108095_107693_104341_106323_103375_106890_107515_108411_108333_107694_108290_108084_108081_107961_108074_108121_107806_107787_108298_107629_107319_107242_100458; pgv_pvi=2169244672; pgv_si=s5109965824; BDRCVFR[feWj1Vr5u3D]=I67x6TjHwwYf0; H_PS_PSSID=20742_1440_19860_12012_20847_20856_20837_20771; bdshare_firstime=1471144779316; Hm_lvt_287705c8d9e2073d13275b18dbd746dc=1471144783; Hm_lpvt_287705c8d9e2073d13275b18dbd746dc=1471144783; wise_device=0; TOPMSG=1471144932-0; BDUSS=NjcVplVGJYQm9QNkR2UEFndTg0QTdWakpwYkdUaFRrSGx0WFdUSjhzbGdjZGRYQVFBQUFBJCQAAAAAAAAAAAEAAACvm74eRWNsaXBzZei0AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGDkr1dg5K9XWk; LONGID=515808175')

        print('Request url : ' + request_url)
        response = requests.get(request_url, headers=headers, cookies=cookies)
        if response.status_code == 200:
            html = response.text.replace('<!--', '').replace('-->', '')
            soup = BeautifulSoup(html, 'html.parser')

            try:
                li_all = soup.find_all('li', class_=' j_thread_list clearfix')
                for li in li_all:
                    href = href_prefix + li.a['href']
                    title = li.a['title']
                    reply = li.find_all('span', class_='threadlist_rep_num center_text')[0].text
                    author = li.find_all('span', class_='tb_icon_author')[0].text
                    createtime = li.find_all('span', class_='pull-right is_show_create_time')[0].text
                    print(
                        'href = ' + href + '; title = ' + title + '; reply = ' + reply + '; author = ' + author + '; createtime = ' + createtime)
                    baidutieba = BaiduTieba(href, reply, title, author, createtime)
                    baidutieba_list.append(baidutieba)
            except:
                pass

        elif response.status_code == 403:
            print('Your requests are denied!')
        else:
            print('ERROR!!')

        pn += 25
    return baidutieba_list


def out_excel(filedir, filename, sheetname, baidutieba_list):
    if not os.path.isdir(filedir):
        os.mkdir(filedir)

    filepath = filedir + os.path.sep + filename + '.xlsx'
    wb = Workbook()
    ws = wb.active
    # ws.print_options.verticalCentered = True
    # ws.print_options.horizontalCentered = True
    ws.title = sheetname
    ws.cell(row=1, column=1).value = '帖子链接'
    ws.cell(row=1, column=2).value = '帖子标题'
    ws.cell(row=1, column=3).value = '回复数量'
    ws.cell(row=1, column=4).value = '主题作者'
    ws.cell(row=1, column=5).value = '创建时间'

    rownum = 2

    for baidutieba in baidutieba_list:
        try:
            ws.cell(row=rownum, column=1).value = baidutieba.href
            ws.cell(row=rownum, column=2).value = baidutieba.title
            ws.cell(row=rownum, column=3).value = baidutieba.reply
            ws.cell(row=rownum, column=4).value = baidutieba.author
            ws.cell(row=rownum, column=5).value = baidutieba.createtime
            rownum += 1
        except:
            pass

    wb.save(filepath)
    print('Excel generates successfully......')


if __name__ == '__main__':
    baidutieba_list = crawl('蒙牛')
    out_excel('/tmp/', '百度贴吧-蒙牛', '蒙牛', baidutieba_list)
