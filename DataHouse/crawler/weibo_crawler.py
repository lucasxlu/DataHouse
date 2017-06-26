"""
A weibo spider to get recently released 100 pages weibo by keyword you are searching
"""
import os
import urllib.parse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


class Weibo:
    def __init__(self, username="", userurl="", content="", like="", repost="", comment="", date_and_device=""):
        self.username = username
        self.userurl = userurl
        self.content = content
        self.like = like
        self.repost = repost
        self.comment = comment
        self.date_and_device = date_and_device


def crawl(keyword):
    weibo_list = []
    pagenum = 1

    while pagenum <= 100:
        request_url = 'http://weibo.cn/search/mblog?hideSearchFrame=&keyword=' + urllib.parse.quote(
            keyword) + '&page=' + str(pagenum)
        print(request_url)
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.106 Safari/537.36',
            'Origin': 'http://weibo.cn',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Host': 'weibo.cn',
            'Upgrade-Insecure-Requests': '1'
        }

        with open('../config/weibo_cookies.txt', mode='rt', encoding='UTF-8') as f:
            cookies_string = ''.join(f.readlines()).strip()
        cookies = dict(cookies_are=cookies_string)
        response = requests.get(request_url, headers=headers, cookies=cookies)
        if response.status_code == 200:
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            print(html)
            divs = soup.find_all('div', class_='c')
            for i in range(3, len(divs) - 2):
                # print(divs[i])
                try:
                    username = divs[i].a.text
                    userurl = divs[i].a['href']
                    content = divs[i].find_all('span', class_='ctt')[0].text
                    all_a = divs[i].find_all('a')

                    for each_a in all_a:
                        if each_a['href'].startswith('http://weibo.cn/attitude'):
                            like = each_a.text
                        elif each_a['href'].startswith('http://weibo.cn/repost'):
                            repost = each_a.text
                        elif each_a['href'].startswith('http://weibo.cn/comment'):
                            comment = each_a.text
                        else:
                            pass
                    # like = all_a[len(all_a) - 4].text
                    # repost = all_a[len(all_a) - 3].text
                    # comment = all_a[len(all_a) - 2].text
                    date_and_device = divs[i].find_all('span', class_='ct')[0].text
                    # date_and_device = divs[i].find_all('span', class_='ct')[0].text.split(" ")
                    # date = date_and_device[len(date_and_device) - 3] + date_and_device[len(date_and_device) - 2]
                    # device = date_and_device[len(date_and_device) - 1]

                    print(
                        'Username : ' + username + '; Userurl : ' + userurl + '; content ' + content + '; like : ' + like + '; repost : ' + repost + '; comment : ' + comment + '; date_and_device : ' + date_and_device)
                    weibo = Weibo(username, userurl, content, like, repost, comment, date_and_device)
                    weibo_list.append(weibo)
                except:
                    pass
        else:
            print('ERROR!' + str(response.status_code))
        # time.sleep(3)
        pagenum += 1

    return weibo_list


def out_excel(filedir, filename, sheetname, weibo_list):
    if not os.path.isdir(filedir):
        os.mkdir(filedir)

    filepath = filedir + os.path.sep + filename + '.xlsx'
    wb = Workbook()
    ws = wb.active
    # ws.print_options.verticalCentered = True
    # ws.print_options.horizontalCentered = True
    ws.title = sheetname
    ws.cell(row=1, column=1).value = '用户名'
    ws.cell(row=1, column=2).value = '用户URL'
    ws.cell(row=1, column=3).value = '微博内容'
    ws.cell(row=1, column=4).value = '赞'
    ws.cell(row=1, column=5).value = '转发'
    ws.cell(row=1, column=6).value = '评论'
    ws.cell(row=1, column=7).value = '日期与来源'

    rownum = 2

    for each_weibo in weibo_list:
        ws.cell(row=rownum, column=1).value = each_weibo.username
        ws.cell(row=rownum, column=2).value = each_weibo.userurl
        ws.cell(row=rownum, column=3).value = each_weibo.content
        ws.cell(row=rownum, column=4).value = each_weibo.like
        ws.cell(row=rownum, column=5).value = each_weibo.repost
        ws.cell(row=rownum, column=6).value = each_weibo.comment
        ws.cell(row=rownum, column=7).value = each_weibo.date_and_device
        rownum += 1

    wb.save(filepath)
    print('Excel generates successfully......')


if __name__ == '__main__':
    weibo_list = crawl('王宝强')
    out_excel('/tmp/', '微博-王宝强', '王宝强', weibo_list)
