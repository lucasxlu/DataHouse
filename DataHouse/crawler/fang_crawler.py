"""
    A web crawler for fang.com
"""
import logging
import re
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

logging.basicConfig(format='%(levelname)s:%(asctime)s:%(message)s  \t', level=logging.INFO)

in_list_house_id = []


class FangHouse(object):
    pass


def get_xiaoqu_info():
    continue_crawl = True
    urls = ['https://m.fang.com/fangjia/?c=pinggu&a=ajaxGetList&city=wuhan&orderby=0&p=%d' % i for i in
            range(1, 101, 1)]
    headers = {
        'referer': 'https://m.fang.com/fangjia/wuhan_list_pinggu/',
        'x-requested-with': 'XMLHttpRequest',
        'user-agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 8_0 like Mac OS X) AppleWebKit/600.1.3 (KHTML, like Gecko) '
                      'Version/8.0 Mobile/12A4345d Safari/600.1.4'
    }
    fang_house_list = []
    for url in urls:
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            logging.info(response.url)
            soup = BeautifulSoup(response.text, 'html5lib')
            li_soups = soup.find_all('li')
            for each_li in li_soups:
                if each_li.a['href'].split('/')[-1].split('.')[0].strip() not in in_list_house_id:
                    fang_house = FangHouse()
                    fang_house.idNo = each_li.a['href'].split('/')[-1].split('.')[0].strip()
                    in_list_house_id.append(fang_house.idNo)
                    fang_house.name = None
                    fang_house.location = None
                    fang_house.price = None
                    fang_house.secondHandHouse = None

                    try:
                        fang_house.name = each_li.find_all('div', class_='info')[0].h3.get_text().strip().encode(
                            'Latin1').decode('utf-8')
                        price_node = each_li.find_all('div', class_='info')[0].p.span.get_text() \
                            .encode('Latin1').decode('utf-8').replace('元/㎡', '').replace('元/平', '').strip()
                        fang_house.price = price_node if price_node != '暂无均价' else None
                        fang_house.location = each_li.find_all('div', class_='info')[0].p.get_text().encode(
                            'Latin1').decode(
                            'utf-8').replace(
                            each_li.find_all('div', class_='info')[0].p.span.get_text().encode('Latin1').decode(
                                'utf-8'),
                            '').strip()
                        second_house = each_li.find_all('p')[-1].get_text().encode('Latin1').decode('utf-8')
                        fang_house.secondHandHouse = re.findall('^二手房\d+套$', second_house)[0].replace('二手房',
                                                                                                      '').replace(
                            '套',
                            '') \
                            if re.search('^二手房\d+套$', second_house) is not None else 0
                    except:
                        pass
                    fang_house_list.append(fang_house)
                else:
                    logging.info('data in this url has been crawled...')
                    continue_crawl = False
            print('Page %s has been crawled successfully~' % url.split('=')[-1])
            time.sleep(2)
        elif response.status_code == 304:
            logging.ERROR('304 forbidden!!')
            # time.sleep(2)

    return fang_house_list


def write_excel(list_, filename):
    EXCEL_DIR = '/home/lucasx/PycharmProjects/DataHouse/DataSet/'
    from DataHouse.crawler.file_helper import mkdirs_if_not_exists
    mkdirs_if_not_exists(EXCEL_DIR)
    wb = Workbook()
    ws = wb.active
    ws.title = "HouseInfo"
    ws.cell(row=1, column=1).value = 'idNo'
    ws.cell(row=1, column=2).value = 'name'
    ws.cell(row=1, column=3).value = 'location'
    ws.cell(row=1, column=4).value = 'price'
    ws.cell(row=1, column=5).value = 'secondHandHouse'

    rownum = 2
    for each_item in list_:
        ws.cell(row=rownum, column=1).value = each_item.idNo
        ws.cell(row=rownum, column=2).value = each_item.name
        ws.cell(row=rownum, column=3).value = each_item.location
        ws.cell(row=rownum, column=4).value = each_item.price
        ws.cell(row=rownum, column=5).value = each_item.secondHandHouse
        rownum += 1
    wb.save(EXCEL_DIR + filename + '.xlsx')
    logging.info('Excel生成成功!')


if __name__ == '__main__':
    fang_house_list = get_xiaoqu_info()
    write_excel(fang_house_list, 'fang')
