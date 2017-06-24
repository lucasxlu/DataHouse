import os
import time
import json

import requests
from lxml import etree
from openpyxl import Workbook

JOBCN_DATA_DIR = '../../DataSet/jobcn/'
base_url = 'http://www.jobcn.com/search/listalljob_servlet.ujson'
user_agent = 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.7 Safari/537.36'


def crawl(jobFunction, jobname, pagenum):
    payload = {'p.keyword': '', 'p.keyword2': '', 'p.keywordType': '', 'p.pageNo': pagenum, 'p.pageSize': 40,
               'p.sortBy': 'postdate', 'p.statistics': 'false', 'p.totalRow': '', 'p.cachePageNo': 1,
               'p.cachePosIds': '', 'p.cachePosUpddates': '', 'p.jobFunction': jobFunction}

    headers = {'user-agent': user_agent, 'Accept': 'application/json, text/javascript, */*; q=0.01',
               'Accept-Encoding': 'gzip, deflate', 'Host': 'www.jobcn.com', 'Origin': 'http://www.jobcn.com',
               'Referer': 'http://www.jobcn.com/search/listalljob.xhtml'}

    r = requests.post(base_url, data=payload, headers=headers)
    if r.status_code == 200:
        print(r.json())

        return r.text
    elif r.status_code == 403:
        print('Access Denied!')


def get_max_page(jobFunction):
    base_url = 'http://www.jobcn.com/search/listalljob_servlet.ujson'

    payload = {'p.keyword': '', 'p.keyword2': '', 'p.keywordType': '', 'p.pageNo': 1, 'p.pageSize': 40,
               'p.sortBy': 'postdate', 'p.statistics': 'false', 'p.totalRow': '', 'p.cachePageNo': 1,
               'p.cachePosIds': '', 'p.cachePosUpddates': '', 'p.jobFunction': jobFunction}

    headers = {'Accept': 'application/json, text/javascript, */*; q=0.01',
               'Accept-Encoding': 'gzip, deflate', 'Host': 'www.jobcn.com', 'Origin': 'http://www.jobcn.com',
               'Referer': 'http://www.jobcn.com/search/listalljob.xhtml'}

    r = requests.post(base_url, data=payload, headers=headers)
    max_page = r.json()['pageCount']
    return max_page


def get_xml_joblist(job_xml_path):
    tree = etree.parse(job_xml_path)
    job_info = {}
    for each_job_node in tree.findall('//job'):
        jobFunction = each_job_node.attrib['jobFunction']
        jobname = each_job_node.text

        job_info[jobFunction] = jobname

    return job_info


def json_file_to_list(each_job_json_dir):
    joblist = []
    for each_job_json in os.listdir(each_job_json_dir):
        with open(each_job_json_dir + os.sep + each_job_json, mode='r', encoding='utf-8') as f:
            for each_line in f.readlines():
                try:
                    json_obj = json.loads(each_line, encoding='utf-8')
                    joblist.append(json_obj)
                except:
                    pass
    return joblist


def write_excel(lists, filename):
    wb = Workbook()
    ws = wb.active
    # ws.print_options.verticalCentered = True
    # ws.print_options.horizontalCentered = True
    ws.title = "职位信息"
    ws.cell(row=1, column=1).value = '职位ID'
    ws.cell(row=1, column=2).value = '职位名称'
    ws.cell(row=1, column=3).value = '所属部门'
    ws.cell(row=1, column=4).value = '公司名称'
    ws.cell(row=1, column=5).value = '薪资待遇'
    ws.cell(row=1, column=6).value = '学历要求'
    ws.cell(row=1, column=7).value = '公司福利'
    ws.cell(row=1, column=8).value = '年龄要求'
    ws.cell(row=1, column=9).value = '工作经验'
    ws.cell(row=1, column=10).value = '招聘数量'
    ws.cell(row=1, column=11).value = '工作地点'
    ws.cell(row=1, column=12).value = '联系邮箱'
    ws.cell(row=1, column=13).value = '联系电话'
    ws.cell(row=1, column=14).value = '公司ID'

    rownum = 2

    for each_item in lists:
        info_list = each_item.get('rows')
        for each_job_info_obj in info_list:
            ws.cell(row=rownum, column=1).value = each_job_info_obj['posId']
            ws.cell(row=rownum, column=2).value = each_job_info_obj['posName']
            ws.cell(row=rownum, column=3).value = each_job_info_obj['deptName']
            ws.cell(row=rownum, column=4).value = each_job_info_obj['comName']
            ws.cell(row=rownum, column=5).value = each_job_info_obj['salaryDesc']
            ws.cell(row=rownum, column=6).value = each_job_info_obj['reqDegree']
            ws.cell(row=rownum, column=7).value = each_job_info_obj['benefitTags']
            ws.cell(row=rownum, column=8).value = each_job_info_obj['ageDesc']
            ws.cell(row=rownum, column=9).value = each_job_info_obj['workYearDesc']
            ws.cell(row=rownum, column=10).value = each_job_info_obj['candidatesNum']
            ws.cell(row=rownum, column=11).value = each_job_info_obj['jobLocation']
            ws.cell(row=rownum, column=12).value = each_job_info_obj['email']
            ws.cell(row=rownum, column=13).value = each_job_info_obj['contactTel']
            ws.cell(row=rownum, column=14).value = each_job_info_obj['comId']
            rownum += 1
    wb.save('./%s.xlsx' % filename)
    print('Excel generates successfully......')


def start():
    job_info = get_xml_joblist('job.xml')
    for key, value in job_info.items():
        '''get the max page number via jobFunction value'''
        max_page = get_max_page(key)
        print('is crawling ' + value + ' \' data......')
        index = 1
        while index <= max_page:
            json_obj = crawl(key, value, index)
            if os.path.exists(os.path.join(JOBCN_DATA_DIR, value)) is False or os.path.isdir(
                    os.path.join(JOBCN_DATA_DIR, value)) is False:
                os.makedirs(os.path.join(JOBCN_DATA_DIR, value))
            with open(os.path.join(JOBCN_DATA_DIR, value + '/%s.json' % str(index)), mode='wt', encoding='utf-8') as f:
                f.write(json_obj)
            index += 1

        print('%s\'s data is finished......' % value)
        time.sleep(2)

    print('start outputting Excel file......')
    for each_job_dir in os.listdir(JOBCN_DATA_DIR):
        joblist = json_file_to_list(os.path.join(JOBCN_DATA_DIR, each_job_dir))
        write_excel(joblist, each_job_dir)


if __name__ == '__main__':
    start()
